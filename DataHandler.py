import csv, os
from openpyxl import load_workbook, Workbook
from requests import get

example_csv = "/Users/nickcsapo/Downloads/SalesJan2009.csv"
example_xlsx = "/Users/nickcsapo/Downloads/Financial Sample.xlsx"

#Utility Function(s). Not to be accessed by user who imports library
#os imports slashes as \\. They need to be a single /
def slashFix(input):
    return input.replace("\\","/")

#Library Functions to be Accessed
#Creates a .csv file from data
def create_csv(location_and_name, data):
    concatenated_data = []
    for row_num in range(len(data)):
        for col_num in range(len(data[row_num])):
            concatenated_data.append(data[row_num][col_num])
    with open(location_and_name, 'w') as csvfile:
        writer = csv.writer(csvfile, quoting=csv.QUOTE_ALL)
        writer.writerow(concatenated_data)

#Creates an excel sheet from data
def create_xlsx(location_and_name, *datas):
    wb = Workbook()
    ws = wb.active
    count = 1
    for data in datas:
        if count != 1:
            ws = wb.create_sheet('sheet'+str(count))
        for row_num in range(len(data)):
            for col_num in range(len(data[row_num])):
                cell = ws.cell(row=row_num+1, column=col_num+1)
                cell.value = data[row_num][col_num]
        count += 1
    wb.save(location_and_name)
    return True

#returns value in a defined cell
def get_cell(data, col_num, row_num):
    return data[col_num-1][row_num-1]

#returns values in a defined column
def get_col(data, col_num, col_max):
    return data[col_num-1::col_max]

#Returns a list of all files within a given folder address
def getFile(location):
    items = []
    dir = os.listdir(location)
    for dirItem in dir:
        dirAddr = os.path.join(location, dirItem)
        if os.path.isfile(dirAddr):
            items.append(dirItem)
    return items

#Returns a list of all files and addresses within a given folder address
def getFileAddr(location):
    items = []
    dir = os.listdir(location)
    for dirItem in dir:
        dirAddr = os.path.join(location, dirItem)
        if os.path.isfile(slashFix(dirAddr)):
            items.append(dirAddr)
    return items

#Lists all folders in a given folder address
def getFolder(location):
    items = []
    dir = os.listdir(location)
    for dirItem in dir:
        dirAddr = os.path.join(location, dirItem)
        if os.path.isdir(dirAddr):
            items.append(dirItem)
    return items

#Lists all folders and addresses witin a given folder address
def getFolderAddr(location):
    items = []
    dir = os.listdir(location)
    for dirItem in dir:
        dirAddr = os.path.join(location, dirItem)
        if os.path.isdir(dirAddr):
            items.append(slashFix(dirAddr))
    return items

#returns the html of a webpage
def get_webpage(address):
    html = get(address).text
    return html

#returns list of each value in csv file
def load_csv(file, delimiter = ',', exclusion_list = []):
    if not isinstance(exclusion_list, list): exclusion_list = [exclusion_list]
    values = []
    with open(file, "r") as imported_file:
        for row in imported_file:
            for value in row.split(delimiter):
                if value not in exclusion_list:
                    values.append(value)
    return values

#returns list of each value in xlsx file
def load_xlsx(file, sheet = 'Sheet1'):
    values = []
    wb = load_workbook(file)
    ws = wb[sheet]
    for row in ws.values:
        for value in row:
            values.append(value)
    return values

#Organizes data based on specified number of columns (steps)
def slice_data(data, step):
    return [data[i::step] for i in range(step)]

#Identical to .split method but leaves delimiter
def split_data(data, delimiter):
    data = data.split(delimiter)
    [element.append(delimiter) for element in data]
    return data

#switches rows with columns
def transpose_data(data):
    transposed_data = []
    for row in range(len(data[0])-1):
        temp = []
        for col in range(len(data)-1):
            temp.append(data[col][row])
        transposed_data.append(temp)
    return transposed_data


print("DataHandler.py loaded")